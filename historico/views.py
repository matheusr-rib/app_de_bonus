from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from historico.models import HistoricoAcao
from django.core.paginator import Paginator
from django.db.models import Q
from campanhas.models import Banco
from usuarios.decorators import tipo_usuario_requerido
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from campanhas.models import Campanha
from datetime import datetime
from django.utils.timezone import localtime


@login_required(login_url='login')
@tipo_usuario_requerido('master')
def historico_listagem(request):
    queryset = HistoricoAcao.objects.select_related('campanha', 'usuario')

    # 🔍 Captura filtros
    campanha_id = request.GET.get('campanha_id')
    nome = request.GET.get('nome')
    banco_nome = request.GET.get('banco')
    usuario_id = request.GET.get('usuario')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    status = request.GET.get('status')

    # ✅ Filtro por ID de campanha (com validação)
    if campanha_id and campanha_id.isdigit():
        queryset = queryset.filter(campanha_id=int(campanha_id))

    # ✅ Filtro por nome usando o campo seguro
    if nome:
        queryset = queryset.filter(campanha_nome__icontains=nome)

    if status:
        queryset = queryset.filter(campanha__status_manual=status)
    if banco_nome:
        queryset = queryset.filter(campanha__banco__nome__icontains=banco_nome)
    if usuario_id:
        queryset = queryset.filter(usuario_id=usuario_id)
    if data_inicio:
        queryset = queryset.filter(vigencia_inicio__gte=data_inicio)
    if data_fim:
        queryset = queryset.filter(vigencia_fim__lte=data_fim)

    # 🔄 Paginação
    paginator = Paginator(queryset.order_by('-data_hora'), 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    bancos = Banco.objects.all()
    usuarios = queryset.values_list('usuario__id', 'usuario__username').distinct()

    status_choices = [
        ('ATIVA', 'Ativa'),
        ('INATIVA', 'Inativa'),
        ('EM ANALISE', 'Em análise'),
    ]

    context = {
        'page_obj': page_obj,
        'bancos': bancos,
        'usuarios': usuarios,
        'status_choices': status_choices,
    }
    return render(request, 'historico_listagem.html', context)


@login_required(login_url='login')
@tipo_usuario_requerido('master')
def relatorios_campanha(request):
    context = {
        'bancos': Campanha.objects.values_list('banco__id', 'banco__nome').distinct(),
        'usuarios': HistoricoAcao.objects.values_list('usuario__id', 'usuario__username').distinct(),
        'status_choices': [
            ('ATIVA', 'Ativa'),
            ('INATIVA', 'Inativa'),
            ('EM ANALISE', 'Em análise'),
        ]
    }
    return render(request, 'relatorios_campanha.html', context)


@login_required(login_url='login')
@tipo_usuario_requerido('master')
def exportar_relatorio_excel(request):
    tipo = request.GET.get("tipo", "alteracoes")
    wb = Workbook()
    ws = wb.active

    if tipo == "vigencias":
        ws.title = "Relatório de Vigências"
        headers = [
            "ID Campanha","Banco", "Campanha", "Status", "Vigência Início", "Vigência Fim",
            "Faixa Inicial", "Faixa Final", "Tipo Valor", "Valor Recebido", "Faixa Garantida"
        ]
        ws.append(headers)

        campanhas = Campanha.objects.select_related("banco").prefetch_related("faixas")

        campanha_id = request.GET.get('campanha_id')
        nome = request.GET.get("nome")
        banco_nome = request.GET.get("banco")
        status = request.GET.get("status")
        data_inicio = request.GET.get("data_inicio")
        data_fim = request.GET.get("data_fim")

        if campanha_id:
            campanhas = campanhas.filter(id=campanha_id)
        if nome:
            campanhas = campanhas.filter(campanha__icontains=nome)
        if banco_nome:
            campanhas = campanhas.filter(banco__nome__icontains=banco_nome)
        if status:
            campanhas = campanhas.filter(status_manual=status)

        if data_inicio and data_fim:
            campanhas = campanhas.filter(
                vigencia_inicio__lte=data_fim
            ).filter(
                Q(vigencia_fim__gte=data_inicio) | Q(vigencia_fim__isnull=True)
            )
        elif data_inicio:
            campanhas = campanhas.filter(
                Q(vigencia_fim__gte=data_inicio) | Q(vigencia_fim__isnull=True)
            )
        elif data_fim:
            campanhas = campanhas.filter(vigencia_inicio__lte=data_fim)

        for c in campanhas:
            vigencia_inicio = c.vigencia_inicio.strftime("%d/%m/%Y") if c.vigencia_inicio else ""
            vigencia_fim = c.vigencia_fim.strftime("%d/%m/%Y") if c.vigencia_fim else "—"

            faixas = c.faixas.all()
            if faixas.exists():
                for faixa in faixas:
                    ws.append([
                        c.id,  # ✅ ID da campanha aqui
                        c.banco.nome,
                        c.campanha,
                        c.status_manual,
                        vigencia_inicio,
                        vigencia_fim,
                        float(faixa.faixa_inicial),
                        float(faixa.faixa_final) if faixa.faixa_final else "Acima",
                        faixa.tipo_valor,
                        float(faixa.valor_recebido),
                        "Sim" if faixa.faixa_garantida else "Não"
                    ])
            else:
                ws.append([
                    c.id,  # ✅ ID da campanha aqui
                    c.banco.nome,
                    c.campanha,
                    c.status_manual,
                    vigencia_inicio,
                    vigencia_fim,
                    "", "", "", "", ""
                ])

    else:
        # Relatório de alterações
        return gerar_relatorio_alteracoes(request, wb, ws)

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"relatorio_{tipo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


def gerar_relatorio_alteracoes(request, wb, ws):
    ws.title = "Relatório de Alterações"
    headers = ["ID Campanha", "Período da Ação", "Banco", "Campanha", "Usuário", "Status"]
    ws.append(headers)

    queryset = HistoricoAcao.objects.select_related("campanha__banco", "usuario")

    campanha_id = request.GET.get('campanha_id')
    nome = request.GET.get('nome')
    banco_nome = request.GET.get('banco')
    usuario_id = request.GET.get('usuario')
    data_inicio = request.GET.get('data_acao_inicio')
    data_fim = request.GET.get('data_acao_fim')
    status = request.GET.get('status')

    if campanha_id:
        queryset = queryset.filter(campanha_id=campanha_id)
    if status:
        queryset = queryset.filter(campanha__status_manual=status)
    if nome:
        queryset = queryset.filter(campanha_nome__icontains=nome)
    if banco_nome:
        queryset = queryset.filter(campanha__banco__nome__icontains=banco_nome)
    if usuario_id:
        queryset = queryset.filter(usuario_id=usuario_id)
    if data_inicio:
        queryset = queryset.filter(data_hora__date__gte=data_inicio)
    if data_fim:
        queryset = queryset.filter(data_hora__date__lte=data_fim)

    queryset = queryset.order_by('-data_hora')

    for h in queryset:
        data_acao = localtime(h.data_hora).strftime("%d/%m/%Y %H:%M")
        banco_nome = h.campanha.banco.nome if h.campanha and h.campanha.banco else "—"
        usuario_nome = h.usuario.username if h.usuario else "Sistema"

        # ✅ Usa método do model para obter nome amigável
        status_display = h.acao_legivel()

        ws.append([
            h.campanha_id or "—",
            data_acao,
            banco_nome,
            h.campanha_nome,
            usuario_nome,
            status_display
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"relatorio_alteracoes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response